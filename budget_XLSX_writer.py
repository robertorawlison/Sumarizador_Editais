# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
import tempfile

def adjust_width_descr(ws, col):
    coluna = ws[col]
    max_length = 0
    coluna = list(coluna)
    for celula in coluna:
        if celula.value:
            descr = str(celula.value)
            for s in descr.split("\n"):
                if len(s) > max_length:
                    max_length = len(s)
        
    ws.column_dimensions[coluna[0].column_letter].width = max_length
    
    for cell in coluna:
        cell.alignment = Alignment(wrap_text=True)

def adjust_width(ws, cols):
    # Ajustar a largura das colunas aos dados
    for col in cols:
        coluna = ws[col]
        max_length = 0
        coluna = list(coluna)
        for celula in coluna:
            if celula.value:
                if len(str(celula.value)) > max_length:
                    max_length = len(celula.value)
        ws.column_dimensions[coluna[0].column_letter].width = max_length

class BudgetXLSXWriter:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = Workbook()
      
    def write(self, budget):
        self.__write_budget(budget.itens_df)
        self.__write_ABC(budget.abc)
        self.__write_benford(budget.benford)
        
        self.wb.save(self.file_name) 
    
    def __write_budget(self, df):
        #Colocando o orçamento na sheet ativa
        ws = self.wb.active
        ws.title = 'Orçamento'
        for row in dataframe_to_rows(df, index=False, header=True):
          ws.append(row)
    
        # Definir a formatação de moeda para a coluna
        for col in ['E', 'F']:
          for cell in ws[col]:
            cell.number_format = '"R$" #,##0.00 '  # Formatação personalizada para moeda (Real brasileiro)
       
        adjust_width(ws, ['C', 'D', 'E', 'F'])
        adjust_width_descr(ws, 'B')
    
        
        
    def __write_ABC(self, abc):
        # Create a new sheet in the existing workbook
        abc_sheet = self.wb.create_sheet('ABC')
    
        # Write the data from abc_df to the new sheet
        for row in dataframe_to_rows(abc.create_df(), index=False, header=True):
            abc_sheet.append(row)
    
        # Definir a formatação de moeda para a coluna
        for col in ['E', 'F', 'H']:
          for cell in abc_sheet[col]:
            cell.number_format = '"R$" #,##0.00 '  # Formatação personalizada para moeda (Real brasileiro)
        
        for col in ['G', 'I']:
          for cell in abc_sheet[col]:
              cell.number_format = '0.0%'
        
        adjust_width(abc_sheet, ['C', 'D', 'E', 'F', 'H'])
        adjust_width_descr(abc_sheet, 'B')
        
        cor_A =  "F0F8FF"
        cor_B = "AED6F1"
        cor_C = "B0C4DE"

        # Colorindo as linhas da tabela (exceto a primeira linha de cabeçalho)
        #Região A
        for row in list(abc_sheet.iter_rows(min_row=2))[0:abc.ponto_a[0]]:
            for cell in row:
                cell.fill = PatternFill(start_color=cor_A, end_color=cor_A, fill_type="solid")
        #Região B
        for row in list(abc_sheet.iter_rows(min_row=2))[abc.ponto_a[0]:abc.ponto_b[0]]:
            for cell in row:
                cell.fill = PatternFill(start_color=cor_B, end_color=cor_B, fill_type="solid")
        #Região C
        for row in list(abc_sheet.iter_rows(min_row=2))[abc.ponto_b[0]:]:
            for cell in row:
                cell.fill = PatternFill(start_color=cor_C, end_color=cor_C, fill_type="solid")
        
        self.__write_figure(abc.get_figure_plot(), "ABC", "l1")
        
        
    def __write_benford(self, benford):
        # Create a new sheet in the existing workbook
        benford_sheet = self.wb.create_sheet('Benford')
    
        # Write the data from abc_df to the new sheet
        for row in dataframe_to_rows(benford.create_df(), index=False, header=True):
            benford_sheet.append(row)
            
        for col in ['C', 'E', 'G']:
          for cell in benford_sheet[col]:
              cell.number_format = '0.0%'
        self.__write_figure(benford.get_figure_plot(), "Benford", "k1")

            
    def __write_figure(self, figure, col, cell_anchor):
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile:
            figure.savefig(tmpfile.name)
            image = Image(tmpfile.name)
            image.anchor = cell_anchor
            benford_sheet = self.wb[col]
            benford_sheet.add_image(image)
    
    
  


